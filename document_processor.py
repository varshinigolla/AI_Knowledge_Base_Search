import os
import uuid
from typing import List, Dict, Any, Optional
from datetime import datetime
import PyPDF2
import docx
import pandas as pd
from openpyxl import load_workbook
import chromadb
from chromadb.config import Settings as ChromaSettings
from sentence_transformers import SentenceTransformer
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.schema import Document
import json

from config import settings
from models import DocumentMetadata

class DocumentProcessor:
    def __init__(self):
        self.embedding_model = SentenceTransformer(settings.EMBEDDING_MODEL)
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.CHUNK_SIZE,
            chunk_overlap=settings.CHUNK_OVERLAP
        )
        
        # Initialize ChromaDB
        self.chroma_client = chromadb.PersistentClient(
            path=settings.CHROMA_PERSIST_DIRECTORY,
            settings=ChromaSettings(anonymized_telemetry=False)
        )
        
        # Create or get collection
        self.collection = self.chroma_client.get_or_create_collection(
            name="knowledge_base",
            metadata={"hnsw:space": "cosine"}
        )
        
        # Ensure upload directory exists
        os.makedirs(settings.UPLOAD_DIRECTORY, exist_ok=True)

    def extract_text_from_file(self, file_path: str, content_type: str) -> str:
        """Extract text from various file formats"""
        try:
            if content_type == "application/pdf":
                return self._extract_pdf_text(file_path)
            elif content_type == "text/plain":
                return self._extract_txt_text(file_path)
            elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                return self._extract_docx_text(file_path)
            elif content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
                return self._extract_excel_text(file_path)
            elif content_type == "text/csv":
                return self._extract_csv_text(file_path)
            else:
                raise ValueError(f"Unsupported file type: {content_type}")
        except Exception as e:
            raise Exception(f"Error extracting text from {file_path}: {str(e)}")

    def _extract_pdf_text(self, file_path: str) -> str:
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text

    def _extract_txt_text(self, file_path: str) -> str:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()

    def _extract_docx_text(self, file_path: str) -> str:
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text

    def _extract_excel_text(self, file_path: str) -> str:
        workbook = load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text

    def _extract_csv_text(self, file_path: str) -> str:
        df = pd.read_csv(file_path)
        return df.to_string()

    def process_document(self, file_path: str, filename: str, content_type: str) -> DocumentMetadata:
        """Process and store a document in the knowledge base"""
        try:
            # Extract text
            text = self.extract_text_from_file(file_path, content_type)
            
            if not text.strip():
                raise ValueError("No text content found in the document")
            
            # Split into chunks
            documents = self.text_splitter.split_text(text)
            
            # Create document ID
            doc_id = str(uuid.uuid4())
            
            # Generate embeddings and store in ChromaDB
            embeddings = self.embedding_model.encode(documents).tolist()
            
            # Prepare metadata for each chunk
            metadatas = []
            ids = []
            
            for i, doc in enumerate(documents):
                chunk_id = f"{doc_id}_chunk_{i}"
                metadata = {
                    "filename": filename,
                    "chunk_index": i,
                    "total_chunks": len(documents),
                    "upload_date": datetime.now().isoformat(),
                    "content_type": content_type
                }
                
                metadatas.append(metadata)
                ids.append(chunk_id)
            
            # Store in ChromaDB
            self.collection.add(
                embeddings=embeddings,
                documents=documents,
                metadatas=metadatas,
                ids=ids
            )
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            return DocumentMetadata(
                filename=filename,
                upload_date=datetime.now().isoformat(),
                file_size=file_size,
                content_type=content_type,
                chunk_count=len(documents),
                processing_status="completed"
            )
            
        except Exception as e:
            return DocumentMetadata(
                filename=filename,
                upload_date=datetime.now().isoformat(),
                file_size=0,
                content_type=content_type,
                chunk_count=0,
                processing_status=f"error: {str(e)}"
            )

    def search_documents(self, query: str, top_k: int = None) -> List[Dict[str, Any]]:
        """Search for relevant document chunks"""
        if top_k is None:
            top_k = settings.TOP_K_RESULTS
            
        # Generate query embedding
        query_embedding = self.embedding_model.encode([query]).tolist()[0]
        
        # Search in ChromaDB
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=top_k,
            include=["documents", "metadatas", "distances"]
        )
        
        # Format results
        formatted_results = []
        for i in range(len(results["documents"][0])):
            formatted_results.append({
                "content": results["documents"][0][i],
                "metadata": results["metadatas"][0][i],
                "similarity_score": 1 - results["distances"][0][i],  # Convert distance to similarity
                "chunk_id": results["ids"][0][i]
            })
        
        return formatted_results

    def get_document_list(self) -> List[DocumentMetadata]:
        """Get list of all processed documents"""
        # This is a simplified version - in production, you'd want to maintain a separate metadata store
        try:
            # Get all documents from ChromaDB
            all_docs = self.collection.get(include=["metadatas"])
            
            # Group by filename and create metadata
            doc_metadata = {}
            for i, metadata in enumerate(all_docs["metadatas"]):
                filename = metadata["filename"]
                if filename not in doc_metadata:
                    doc_metadata[filename] = DocumentMetadata(
                        filename=filename,
                        upload_date=metadata["upload_date"],
                        content_type=metadata["content_type"],
                        chunk_count=0,
                        file_size=0,  # We don't store this in ChromaDB
                        processing_status="completed"
                    )
                doc_metadata[filename].chunk_count += 1
            
            return list(doc_metadata.values())
        except Exception as e:
            return []

    def delete_document(self, filename: str) -> bool:
        """Delete a document and all its chunks from the knowledge base"""
        try:
            # Get all chunks for this document
            all_docs = self.collection.get(include=["metadatas", "ids"])
            
            chunk_ids_to_delete = []
            for i, metadata in enumerate(all_docs["metadatas"]):
                if metadata["filename"] == filename:
                    chunk_ids_to_delete.append(all_docs["ids"][i])
            
            if chunk_ids_to_delete:
                self.collection.delete(ids=chunk_ids_to_delete)
                return True
            return False
        except Exception as e:
            return False
